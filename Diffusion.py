# -*- coding: utf-8 -*-
"""
Created on Thu Oct 24 15:46:20 2019

@author: wjlee
"""

import networkx as nx

Result = {}
Result['DFOA'] = {}
Result['CDIOA'] = {}
Result['CDFOA'] = {}

def SaveResult(alg, var):
    Result[alg]['q'] = var[0]
    Result[alg]['c'] = var[1]
    Result[alg]['총 node수'] = var[2]
    Result[alg]['Overlap node수'] = var[3]
    Result[alg]['1hop neighbor node수'] = var[4]
    return

def initialize(GraphName = ""):
    import os
    print("........ Read the graph")
    GRAPH = os.path.join(GraphName, GraphName + 'Graph.gpickle')
    G = nx.read_gpickle(GRAPH)
    
    print("........ Set behavior of all nodes B")
    for node in list(G.nodes):
        G.nodes[node]['behavior'] = 'B'
    
    print("........ Save overlapped nodes")
    InOverlap = []
    for node in list(G.nodes):
        if len(G.nodes[node]['community']) >= 2:
            InOverlap.append(node)
    
    import numpy as np
    # Permute InOverlap
    order = np.random.permutation(len(InOverlap))
    InOverlap = list(np.array(InOverlap)[order])
    
    print("........ Save 1hop neighbors of overlapped nodes")
    _1hop = []
    for node in G.nodes():
        if set(G.neighbors(node)) & set(InOverlap) and node not in InOverlap:
            _1hop.append(node)
            _1hop = list(set(_1hop))
    
    # Permute _1hop
    order = np.random.permutation(len(_1hop))
    _1hop = list(np.array(_1hop)[order])
    
    return G, InOverlap, _1hop

def DFOA(GraphName="", H=None, InOverlap=None, _1hop=None, q=0):
    print("........ * Diffusion from Overlapped Area - Behavior: A, B")
    
    # If q value is 0, then select q value randomly
    if q == 0:
        import numpy as np
        print("........ Randomly choose q value")
        q = np.random.uniform(0, 0.5)
    
    GRAPH = H.copy()
    NumA, NumB = [], []
    print("........ * Threshold q : {0:.4f}".format(q))
    
    # Initialize the behaviors of nodes in overlapped region
    for node in InOverlap:
        GRAPH.nodes[node]['behavior'] = 'A'
    
    # Step 0
    Beh_A = len([x for x in GRAPH.nodes() if GRAPH.nodes[x]['behavior'] == 'A'])
    Beh_B = len([x for x in GRAPH.nodes() if GRAPH.nodes[x]['behavior'] == 'B'])
    NumA.append(Beh_A)
    NumB.append(Beh_B)
    
    # Diffusion
    step = 0
    print("........ STEP : {0} Behavior A : {1} Behavior B : {2} Total : {3}".format(step, Beh_A, Beh_B, len(GRAPH.nodes())))

    # Save information of Graph in step 0
    var = [q, 'x', len(list(GRAPH.nodes)), len(InOverlap), len(_1hop)]
    SaveResult('DFOA', var)
    
    Looked = InOverlap.copy()
    while step < 100:
        step += 1
        ToSee = []  # The nodes who are 1hop neighbor of infected nodes
                    # and they haven't changed their behavior
        hop1 = []   # 1hop neighbors of nodes who already changed their behavior
        for node in Looked:
            tmp = list(GRAPH.neighbors(node))
            hop1 = list(set(hop1) | set(tmp))
        ToSee = list(set(ToSee) | (set(hop1) - set(Looked)))
        
        changedNode = []
        for v in ToSee:
            # NvtoA: Ratio of nodes who have behavior A
            NvtoA = len([n for n in GRAPH.neighbors(v) if GRAPH.nodes[n]['behavior'] == 'A'])/GRAPH.degree(v)
            # Condition for changing behavior
            if NvtoA >= q:
                changedNode.append(v)
        # if there is no node who will change behavior, then stop
        if not changedNode: break
        for v in changedNode:
            GRAPH.nodes[v]['behavior'] = 'A'
        
        Beh_A = len([x for x in GRAPH.nodes() if GRAPH.nodes[x]['behavior'] == 'A'])
        Beh_B = len([x for x in GRAPH.nodes() if GRAPH.nodes[x]['behavior'] == 'B'])
        NumA.append(Beh_A)
        NumB.append(Beh_B)
        
        print("........ STEP : {0} Behavior A : {1} Behavior B : {2} Total : {3}".format(step, Beh_A, Beh_B, len(GRAPH.nodes())))
        
        tmp = [node for node in changedNode]
        Looked = list(set(tmp) | set(Looked))
    
    # Save the result
    Result['DFOA']['A'] = {}
    Result['DFOA']['B'] = {}
    for idx, num in enumerate(NumA):
        Result['DFOA']['A']['step '+str(idx)] = num
    for idx, num in enumerate(NumB):
        Result['DFOA']['B']['step '+str(idx)] = num
    SaveResultToExcel(GraphName, 'DFOA', Result, q, -1)
    return

def CDIOA(GraphName="", H=None, InOverlap=None, _1hop=None, q=0, c=0):
    
    import math
    import numpy as np
    print("........ * Compatibility Diffusion Into Overlapped Area - Behavior: A, B, AB")
    if q == 0 and c == 0:
        # Randomly choose the values of q and c
        # q and c value satisfy the condition for diffusion
        print("........ Randomly Choose q and c value")
        c = np.random.uniform(0, 1/4)
        Min = math.ceil((1-math.sqrt(1-4*c))/2 * 1000)/1000
        Max = 1/2
        q = np.random.uniform(Min, Max)
    
    GRAPH = H.copy()
    
    print("........ * Threshold q : {0:.4f}, Cost c : {1:.4f}".format(q,c))
    
    # Initialize behaviors
    Behaviors = ['A', 'B', 'AB']
    print("........ * Initialize the behaviors of 1-hop neighbors of overlapped region")
    for idx, node in enumerate(_1hop):
        if idx < len(_1hop)/2: GRAPH.nodes[node]['behavior'] = 'A'
        else: GRAPH.nodes[node]['behavior'] = 'AB'
    
    NumA, NumB, NumAB = [], [], []

    # Diffusion
    step = 0
    Beh_A = len([n for n in _1hop if GRAPH.nodes[n]['behavior'] == 'A'])
    Beh_AB = len([n for n in _1hop if GRAPH.nodes[n]['behavior'] == 'AB'])
    Beh_B = len([n for n in GRAPH.nodes() if GRAPH.nodes[n]['behavior'] == 'B'])
    print("........ STEP : {0} Behavior A : {1} Behavior B : {2} Behavior AB : {3} Total : {4}".format(step, Beh_A, Beh_B, Beh_AB, len(GRAPH.nodes())))
    NumA.append(Beh_A)
    NumB.append(Beh_B)
    NumAB.append(Beh_AB)
    
    # Save information of Graph in step 0
    var = [q, c, len(list(GRAPH.nodes)), len(InOverlap), len(_1hop)]
    SaveResult('CDIOA', var)
    
    # Looked: The nodes who already changed their behavior
    Looked = _1hop.copy()
    while step < 100:
        step += 1
        ToSee = [] # The nodes who are 1hop neighbor of infected nodes
                   # and they haven't changed their behavior
        hop1 = []  # 1hop neighbors of nodes who already changed their behavior
        for node in Looked:
            tmp = list(GRAPH.neighbors(node))
            hop1 = list(set(hop1) | set(tmp))
        ToSee = list(set(ToSee) | (set(hop1) - set(Looked)))
        
        changedNode = []
        for v in ToSee:
            # Nvto'behavior': Ratio of nodes who have behavior 'behavior'
            NvtoA = len([n for n in GRAPH.neighbors(v) if GRAPH.nodes[n]['behavior'] == 'A'])/GRAPH.degree(v)
            NvtoB = len([n for n in GRAPH.neighbors(v) if GRAPH.nodes[n]['behavior'] == 'B'])/GRAPH.degree(v)
            NvtoAB = len([n for n in GRAPH.neighbors(v) if GRAPH.nodes[n]['behavior'] == 'AB'])/GRAPH.degree(v)
            
            # Payoff from choosing behaviors
            P_A = (1-q)*(NvtoA + NvtoAB)
            P_B = q*(NvtoB + NvtoAB)
            P_AB = P_A + q*NvtoB - c
            
            # Select behavior of the highest payoff
            Behav = GRAPH.nodes[v]['behavior']
            candidate = np.array([P_A, P_B, P_AB])
            if Behav != Behaviors[np.argmax(candidate)]:
                changedNode.append([v, Behaviors[np.argmax(candidate)]])
        # if there is no node who will change behavior, then stop
        if not changedNode: break
        for node, behavior in changedNode:
            GRAPH.nodes[node]['behavior'] = behavior
            
        Beh_A = len([n for n in GRAPH.nodes() if GRAPH.nodes[n]['behavior'] == 'A'])
        Beh_B = len([n for n in GRAPH.nodes() if GRAPH.nodes[n]['behavior'] == 'B'])
        Beh_AB = len([n for n in GRAPH.nodes() if GRAPH.nodes[n]['behavior'] == 'AB'])
        NumA.append(Beh_A)
        NumB.append(Beh_B)
        NumAB.append(Beh_AB)
        print("........ STEP : {0} Behavior A : {1} Behavior B : {2} Behavior AB : {3} Total : {4}".format(step, Beh_A, Beh_B, Beh_AB, len(GRAPH.nodes())))
        
        tmp = [node for node, behav in changedNode]
        Looked =  list(set(tmp) | set(Looked))
    
    # Save the result
    Result['CDIOA']['A'] = {}
    Result['CDIOA']['B'] = {}
    Result['CDIOA']['AB'] = {}
    Result['CDIOA']['1hop neighbor node 수'] = len(_1hop)
    for idx, num in enumerate(NumA):
        Result['CDIOA']['A']['step '+str(idx)] = num
    for idx, num in enumerate(NumB):
        Result['CDIOA']['B']['step '+str(idx)] = num
    for idx, num in enumerate(NumAB):
        Result['CDIOA']['AB']['step '+str(idx)] = num
        
    SaveResultToExcel(GraphName, 'CDIOA', Result, q, c)
    
    return

def CDFOA(GraphName="", H=None, InOverlap=None, _1hop=None, q=0, c=0):
    import math
    import numpy as np
    
    print("........ * Diffusion from Overlapped Area - Behavior: A, B, AB")
    if q == 0 and c == 0:
        # Randomly choose the value of q and c
        # q and c value satify the condition for diffusion
        print("........ Randomly Choose q and c value")
        c = np.random.uniform(0, 1/4)
        Min = math.ceil((1-math.sqrt(1-4*c))/2 * 1000)/1000
        Max = 1/2
        q = np.random.uniform(Min, Max)
        
    GRAPH = H.copy()
    
    print("........ * Threshold q : {0:.4f}, Cost c : {1:.4f}".format(q,c))
    
    # Initialize the behaviors
    print("........ Initialize the behaviors in Overlapped region")
    Behaviors = ['A', 'B', 'AB']

    for idx, node in enumerate(InOverlap):
        if idx < len(InOverlap)/2: GRAPH.nodes[node]['behavior'] = 'A'
        else: GRAPH.nodes[node]['behavior'] = 'AB'
    
    NumA, NumB, NumAB = [], [], []
    
    # Diffusion
    step = 0
    Beh_A = len([n for n in InOverlap if GRAPH.nodes[n]['behavior'] == 'A'])
    Beh_AB = len([n for n in InOverlap if GRAPH.nodes[n]['behavior'] == 'AB'])
    Beh_B = len([n for n in GRAPH.nodes() if GRAPH.nodes[n]['behavior'] == 'B'])
    print("........ STEP : {0} Behavior A : {1} Behavior B : {2} Behavior AB : {3} Total : {4}".format(step, Beh_A, Beh_B, Beh_AB, len(GRAPH.nodes())))
    NumA.append(Beh_A)
    NumB.append(Beh_B)
    NumAB.append(Beh_AB)
    
    # Save information of Graph in step 0
    var = [q, c, len(list(GRAPH.nodes)), len(InOverlap), len(_1hop)]
    SaveResult('CDFOA', var)
    Result['CDFOA']['1hop neighbor node 수'] = len(_1hop)
    
    # Looked: The nodes who already changed their behavior
    Looked = InOverlap.copy()
    while step < 100:
        step += 1
        ToSee = [] # The nodes who are 1hop neighbor of infected nodes
                   # and they haven't changed their behavior
        hop1 = []  # 1hop neighbors of nodes who already changed their behavior
        for node in Looked:
            tmp = list(GRAPH.neighbors(node))
            hop1 = list(set(hop1) | set(tmp))
        ToSee = list(set(ToSee) | (set(hop1) - set(Looked)))

        changedNode = []
        for v in ToSee:
            # Nvto'behavior': Ratio of nodes who have behavior 'behavior'
            NvtoA = len([n for n in GRAPH.neighbors(v) if GRAPH.nodes[n]['behavior'] == 'A']) / GRAPH.degree(v)
            NvtoB = len([n for n in GRAPH.neighbors(v) if GRAPH.nodes[n]['behavior'] == 'B']) / GRAPH.degree(v)
            NvtoAB = len([n for n in GRAPH.neighbors(v) if GRAPH.nodes[n]['behavior'] == 'AB']) / GRAPH.degree(v)
            
            # Payoff from choosing behaviors
            P_A = (1-q)*(NvtoA + NvtoAB)
            P_B = q*(NvtoB + NvtoAB)
            P_AB = P_A + q*NvtoB - c
            
            # Select behavior of the highest payoff
            Behav = GRAPH.nodes[v]['behavior']
            candidate = np.array([P_A, P_B, P_AB])
            if Behav != Behaviors[np.argmax(candidate)]:
                changedNode.append([v, Behaviors[np.argmax(candidate)]])
        # if there is no node who will change behavior, then stop
        if not changedNode: break
        for node, behavior in changedNode:
            GRAPH.nodes[node]['behavior'] = behavior
        
        Beh_A = len([n for n in GRAPH.nodes() if GRAPH.nodes[n]['behavior'] == 'A'])
        Beh_B = len([n for n in GRAPH.nodes() if GRAPH.nodes[n]['behavior'] == 'B'])
        Beh_AB = len([n for n in GRAPH.nodes() if GRAPH.nodes[n]['behavior'] == 'AB'])
        NumA.append(Beh_A)
        NumB.append(Beh_B)
        NumAB.append(Beh_AB)
        print("........ STEP : {0} Behavior A : {1} Behavior B : {2} Behavior AB : {3} Total : {4}".format(step, Beh_A, Beh_B, Beh_AB, len(GRAPH.nodes())))
        
        tmp = [node for node, behav in changedNode]
        Looked = list(set(tmp) | set(Looked))
    
    # Save the result
    Result['CDFOA']['A'] = {}
    Result['CDFOA']['B'] = {}
    Result['CDFOA']['AB'] = {}
    for idx, num in enumerate(NumA):
        Result['CDFOA']['A']['step '+str(idx)] = num
    for idx, num in enumerate(NumB):
        Result['CDFOA']['B']['step '+str(idx)] = num
    for idx, num in enumerate(NumAB):
        Result['CDFOA']['AB']['step '+str(idx)] = num
    
    SaveResultToExcel(GraphName, 'CDFOA', Result, q, c)
    return


def SaveResultToExcel(GraphName, alg, Result, q, c):
    import os
    import openpyxl
    
    filedir = '/' + GraphName + '/Result/'
    if not os.path.exists(filedir[1:]):
        os.makedirs(filedir[1:])
        
    if c < 0:
        filename = alg + '_q' + str(round(q,3)) + '.xlsx'
    if c >= 0:
        filename = alg + '_q' + str(round(q,3)) + '_c' + str(round(c,3)) + '.xlsx'
    
    if filename in os.listdir(os.getcwd() + '/' + filedir):
        wb = openpyxl.load_workbook(filedir + filename)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.get_sheet_by_name('Sheet'))
    
    ws = wb.create_sheet(alg)
    n_row = 1
    for key, value in Result[alg].items():
        if type(value) != dict:
            ws.cell(column = 1, row = n_row, value = key)
            ws.cell(column = 2, row = n_row, value = value)
        else:
            for _key, _value in Result[alg][key].items():
                ws.cell(column = 1, row = n_row, value = key)
                ws.cell(column = 2, row = n_row, value = _key)
                ws.cell(column = 3, row = n_row, value = _value)
                n_row+=1
        n_row += 1
    
    wb.save(os.getcwd() + filedir + filename)
    wb.close()
